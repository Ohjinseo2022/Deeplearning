from openpyxl import load_workbook  # 엑셀파일을 불러오기 위한 라이브러리

# load_workbook 에서 사용할수 있는 속성값
# -filename
# -read_only  읽기만 해서 쓰기제한을 둘것인지 설정하는 파라미터입니다. 아무 언급을 안하면 쓰기도 가능한 상태로 엑셀을 불러들입니다.
# -keep_vba 엑셀에 있는 vba(액셀의 매크로기능 Visual Basic for Applications) 기능을 살려서 가져올것인지 아닌지를 설정합니다. 기본적으로 엑셀에 저장되어 있는 vba은 가져오지 않는 것으로 설정되어있습니다
# -data_only  엑셀의 수식값을 가져올것인지 수식의 결과인 data를 가져올지를 묻는 설정입니다.
# -keep_links 엑셀파일안에 있는 링크를 그대로 쓸것인지를 설정할 수 있습니다.
load_wb = load_workbook(r"수료증명단.xlsx")
load_ws = load_wb.active  # 가져온 엑셀파일 가져오기

name_list = []
birthday_list = []
ho_list = []

##반복문을 통하여 각각의 리스트에 이름, 생일, 번호 순서대로 접근하여 추가
for i in range(1, load_ws.max_row + 1):  # max_row -> 입력된 행의 갯수를 받아옴
    name_list.append(load_ws.cell(i, 1).value)  # cell -> 각각의 행과 열에 접근하여 값을 가져오는 함수
    birthday_list.append(load_ws.cell(i, 2).value)
    ho_list.append(load_ws.cell(i, 3).value)

print(name_list)
print(birthday_list)
print(ho_list)
